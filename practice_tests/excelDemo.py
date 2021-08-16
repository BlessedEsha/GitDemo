import openpyxl
workbook = openpyxl.load_workbook("C:\\ExcelDemo.xlsx",data_only=True)
sheet = workbook.active
print(sheet)
print(sheet.cell(row=1,column=1).value)

#sheet["C16"].value = "Update"
#workbook.save("C:\\ExcelDemo.xlsx")
a=""
dic = {}
dic_copy = {}
l = []
for i in range(1,sheet.max_row):
    for j in range(1,sheet.max_column+1):
        dic[sheet.cell(row=1, column=j).value]=sheet.cell(row=i+1, column=j).value
        a = a + sheet.cell(row=i, column=j).value + " "

   # print(dic)
    dic_copy = dic.copy()
    print(dic_copy)
    l.append(dic_copy)
    print(l[0]["firstname"])
        # print(sheet.cell(row=i,column=j).value)
    #
    print(l)


