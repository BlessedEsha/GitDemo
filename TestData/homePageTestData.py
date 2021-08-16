class HomePageTestData:
    test_homePage_Data = [{"firstname" : "Hitesh", "email" : "hitesh.dutta@gmail.com" , "gender" : "Male"},
                            {"firstname" : "esha", "email" : "ereshasharma@gmail.com" , "gender" : "Female"}]
    @staticmethod
    def get_testData(test_case_name):
        import openpyxl
        workbook = openpyxl.load_workbook("C:\\ExcelDemo.xlsx",data_only=True)
        sheet = workbook.active
        print(sheet)
        print(sheet.cell(row=1, column=1).value)

        # sheet["C16"].value = "Update"
        # workbook.save("C:\\ExcelDemo.xlsx")
        a = ""
        dic = {}
        dic_copy = {}
        l = []
        for i in range(1, sheet.max_row):
            if sheet.cell(row=i, column=1).value == test_case_name:
                print(test_case_name)

                for j in range(2, sheet.max_column + 1):
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    a = a + sheet.cell(row=i, column=j).value + " "



        return[dic]
        #HomePageTestData.get_testData("TestCaseName3")